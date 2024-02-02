import common
import quote

import openpyxl

class WeeklyQuotes:

    def __init__(self, brand) -> None:
        self.__brand = brand
        self.__weekly_quotes_list = list()

    def append(self, period, open, high, low, close, volume):
        self.__weekly_quotes_list.append(
            quote.Quote(period, open, high, low, close, volume))

    def brand(self):
        return self.__brand

    def list(self):
        return self.__weekly_quotes_list

    def write_xlsx_header(self, worksheet, row):
        worksheet[f'A{row}'] = 'week'
        worksheet[f'B{row}'] = 'open'
        worksheet[f'C{row}'] = 'high'
        worksheet[f'D{row}'] = 'low'
        worksheet[f'E{row}'] = 'close'
        worksheet[f'F{row}'] = 'volume'

    def write_xslx(self, workbook, begin_row):
        __worksheet = workbook.create_sheet(title="weekly")
        self.write_xlsx_header(__worksheet, begin_row)
        for __weekly_quotes_index, __weekly_quotes in enumerate(self.__weekly_quotes_list, begin_row + 1):
            __worksheet[f'A{__weekly_quotes_index}'] = __weekly_quotes.period()
            __worksheet[f'B{__weekly_quotes_index}'] = __weekly_quotes.open()
            __worksheet[f'C{__weekly_quotes_index}'] = __weekly_quotes.high()
            __worksheet[f'D{__weekly_quotes_index}'] = __weekly_quotes.low()
            __worksheet[f'E{__weekly_quotes_index}'] = __weekly_quotes.close()
            __worksheet[f'F{__weekly_quotes_index}'] = __weekly_quotes.volume()

        __candle_stick_chart = openpyxl.chart.StockChart()
        __candle_stick_labels = openpyxl.chart.Reference(__worksheet,min_col=1,min_row=2,max_row=__worksheet.max_row)
        __candle_stick_data = openpyxl.chart.Reference(__worksheet,min_col=2,max_col=5,min_row=2,max_row=__worksheet.max_row)
        __candle_stick_chart.add_data(__candle_stick_data,titles_from_data=True)
        for __series in __candle_stick_chart.series:
            __series.graphicalProperties.line.noFill = True
        __candle_stick_chart.hiLowLines = openpyxl.chart.axis.ChartLines()
        __candle_stick_chart.upDownBars = openpyxl.chart.updown_bars.UpDownBars()

        __candle_stick_pts = [openpyxl.chart.data_source.NumVal(idx=i) for i in range(len(__candle_stick_data) - 1)]
        __candle_stick_cache = openpyxl.chart.data_source.NumData(pt=__candle_stick_pts)
        __candle_stick_chart.series[-1].val.numRef.numCache = __candle_stick_cache

        __candle_stick_chart.set_categories(__candle_stick_labels)
        __candle_stick_chart.width = 40
        __candle_stick_chart.height = 23
        __candle_stick_chart.legend = None
        
        __candle_stick_graph = workbook.create_sheet(title='weekly chart')
        __candle_stick_graph.add_chart(__candle_stick_chart,"A1")

    def calc(daily_quotes):
        __period = '1900-01'
        __brand = daily_quotes.brand()
        __weekly_quotes = WeeklyQuotes(__brand)
        __last_index = len(daily_quotes.list()) - 1

        for __quotes_index, __quotes in enumerate(daily_quotes.list()):
            __quotes_period = common.year_week_number(__quotes.period())
            if __quotes_index == 0:
                __period = __quotes_period
                __open = __quotes.open()
                __high = __quotes.high()
                __low = __quotes.low()
                __volume = __quotes.volume()

            if __period != __quotes_period:
                __weekly_quotes.append(
                    __period, __open, __high, __low, __close, __volume)
                __period = __quotes_period                
                __open = __quotes.open()
                __high = __quotes.high()
                __low = __quotes.low()
                __volume = __quotes.volume()

            if __period == __quotes_period:
                __high = max(__high, __quotes.high())
                __low = min(__low, __quotes.low())
                __volume += __quotes.volume()
                __close = __quotes.close()

            if __quotes_index == __last_index:               
                __high = max(__high, __quotes.high())
                __low = min(__low, __quotes.low())
                __volume += __quotes.volume()
                __close = __quotes.close()
                __weekly_quotes.append(
                    __period, __open, __high, __low, __close, __volume)

        return __weekly_quotes
