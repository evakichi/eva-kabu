import common
import openpyxl
import quote


class QuoteList:

    def __init__(self, brand) -> None:
        self.__brand = brand
        self.__quotes_list = list()

    def append(self, period, open, high, low, close, volume):
        self.__quotes_list.append(
            quote.Quote(period, open, high, low, close, volume))

    def brand(self):
        return self.__brand

    def list(self):
        return self.__quotes_list

    def write_xslx(self, workbook, title, start_row):
        __worksheet = workbook.create_sheet(title=title)

        __worksheet[f'A{start_row}'] = 'month'
        __worksheet[f'B{start_row}'] = 'open'
        __worksheet[f'C{start_row}'] = 'high'
        __worksheet[f'D{start_row}'] = 'low'
        __worksheet[f'E{start_row}'] = 'close'
        __worksheet[f'F{start_row}'] = 'volume'

        for __quotes_index, __quotes in enumerate(self.__quotes_list, start_row + 1):
            __worksheet[f'A{__quotes_index}'] = __quotes.period()
            __worksheet[f'B{__quotes_index}'] = __quotes.open()
            __worksheet[f'C{__quotes_index}'] = __quotes.high()
            __worksheet[f'D{__quotes_index}'] = __quotes.low()
            __worksheet[f'E{__quotes_index}'] = __quotes.close()
            __worksheet[f'F{__quotes_index}'] = __quotes.volume()

        __candle_stick_chart = openpyxl.chart.StockChart()
        __candle_stick_labels = openpyxl.chart.Reference(
            __worksheet, min_col=1, min_row=2, max_row=__worksheet.max_row)
        __candle_stick_data = openpyxl.chart.Reference(
            __worksheet, min_col=2, max_col=5, min_row=2, max_row=__worksheet.max_row)
        __candle_stick_chart.add_data(
            __candle_stick_data, titles_from_data=True)
        for __series in __candle_stick_chart.series:
            __series.graphicalProperties.line.noFill = True
        __candle_stick_chart.hiLowLines = openpyxl.chart.axis.ChartLines()
        __candle_stick_chart.upDownBars = openpyxl.chart.updown_bars.UpDownBars()

        __candle_stick_pts = [openpyxl.chart.data_source.NumVal(
            idx=i) for i in range(len(__candle_stick_data) - 1)]
        __candle_stick_cache = openpyxl.chart.data_source.NumData(
            pt=__candle_stick_pts)
        __candle_stick_chart.series[-1].val.numRef.numCache = __candle_stick_cache

        __candle_stick_chart.set_categories(__candle_stick_labels)
        __candle_stick_chart.width = 40
        __candle_stick_chart.height = 23
        __candle_stick_chart.legend = None

        __volume_chart = openpyxl.chart.BarChart()
        __volume_chart_data = openpyxl.chart.Reference(__worksheet, min_col=6, max_col=6, min_row=2, max_row=__worksheet.max_row)
        
        __volume_chart.add_data(__volume_chart_data, titles_from_data=True)
        __volume_chart.y_axis.axId = 200
        __volume_chart.y_axis.crosses = 'max'

        __candle_stick_chart += __volume_chart

        __candle_stick_graph = workbook.create_sheet(title=title+' chart')
        __candle_stick_graph.add_chart(__candle_stick_chart, "A1")

    def write_xslx25(self, workbook, title):
        __worksheet = workbook[title]
        __candle_stick_chart = openpyxl.chart.StockChart()
        __candle_stick_labels = openpyxl.chart.Reference(
            __worksheet, min_col=1, min_row = max(2,__worksheet.max_row-25) ,max_row=__worksheet.max_row)
        __candle_stick_data = openpyxl.chart.Reference(
            __worksheet, min_col=2, max_col=5, min_row = max(2,__worksheet.max_row-25), max_row=__worksheet.max_row)
        __candle_stick_chart.add_data(
            __candle_stick_data, titles_from_data=True)
        for __series in __candle_stick_chart.series:
            __series.graphicalProperties.line.noFill = True
        __candle_stick_chart.hiLowLines = openpyxl.chart.axis.ChartLines()
        __candle_stick_chart.upDownBars = openpyxl.chart.updown_bars.UpDownBars()

        __candle_stick_pts = [openpyxl.chart.data_source.NumVal(
            idx=i) for i in range(len(__candle_stick_data) - 1)]
        __candle_stick_cache = openpyxl.chart.data_source.NumData(
            pt=__candle_stick_pts)
        __candle_stick_chart.series[-1].val.numRef.numCache = __candle_stick_cache

        __candle_stick_chart.set_categories(__candle_stick_labels)
        __candle_stick_chart.width = 40
        __candle_stick_chart.height = 23
        __candle_stick_chart.legend = None

        __volume_chart = openpyxl.chart.BarChart()
        __volume_chart_data = openpyxl.chart.Reference(__worksheet, min_col=6, max_col=6, min_row = max(2,__worksheet.max_row-25), max_row=__worksheet.max_row)
        
        __volume_chart.add_data(__volume_chart_data, titles_from_data=True)
        __volume_chart.y_axis.axId = 200
        __volume_chart.y_axis.crosses = 'max'

        __candle_stick_chart += __volume_chart

        __candle_stick_graph = workbook.create_sheet(title=title+' chart (past 25 periods)')
        __candle_stick_graph.add_chart(__candle_stick_chart, "A1")
